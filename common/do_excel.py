from openpyxl import load_workbook
from common.read_config import ReadConfig
from common import project_path
from common.project_path import case_path


class DoExcel:
    """对excel表格进行读写操作一级测试结果的写入"""

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name  # Excel工作簿文件名或者地址
        self.sheet_name = sheet_name  # 表单名

    def read_data(self):
        """读取excel内容，有返回值"""
        case_id = ReadConfig(project_path.conf_path).get_data('CASE', 'case_id')
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(2, sheet.max_row+1):
            row_data = {}
            row_data['CaseID'] = sheet.cell(i, 1).value
            row_data['Module'] = sheet.cell(i, 2).value
            row_data['Title'] = sheet.cell(i, 3).value
            row_data['Method'] = sheet.cell(i, 4).value
            row_data['Url'] = sheet.cell(i, 5).value
            row_data['Params'] = sheet.cell(i, 6).value
            row_data['ExpectedResult'] = sheet.cell(i, 7).value
            test_data.append(row_data)
        wb.close()
        final_data = []  # 存储最终的测试数据
        if case_id == 'all':  # 如果等于成立，获取所有的用例数据
            final_data = test_data
        else:  # 等于all不成立，就获取指定case_id的数据
            for i in case_id:
                final_data.append(test_data[i-1])
        return final_data

    def write_back(self, row, column, value):
        """把测试结果写入excel"""
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, column).value = value
        wb.save(self.file_name)
        wb.close()

    # def new_excel(self, sheet_name, file_name):
    #     """新建excel"""
    #     wb = workbook.Workbook()
    #     wb.create_sheet(sheet_name)  # 新建表单
    #     wb.save(file_name)


if __name__ == '__main__':
    test_data = DoExcel(case_path, 'register').read_data()
    print(test_data)